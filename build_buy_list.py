"""
KSI Weekly Buy List Builder
===========================
Rerun weekly after dropping the new raw export(s) into this folder:

    python build_buy_list.py

Inputs (this folder):
  - CAP Raw.xlsx        Northeast location-level YTD export (required)
  - FL Raw.xlsx         Florida location-level export (optional)
  - KSI_Item_master.csv item master (patent flag, velocity, vendors)
  - Vendor and Type.csv vendor -> Domestic / Oversea

Outputs (output/ subfolder):
  - Buy List <date>.xlsx   (Buy List, Vendor Summary, Warehouse Summary,
                            Exceptions, Assumptions)
  - CAP.html               self-contained interactive report (GitHub-ready)
  - index.html             redirect to CAP.html for the Pages homepage

Methodology
-----------
Demand (units/selling-day) = base ADU x seasonal index, per region:
  - Base ADU: 70% last-35-day ADU + 30% YTD ADU when the region's 35-day
    column is a true per-selling-day rate (auto-detected per region);
    otherwise YTD ADU alone. Both exports qualify since the 2026-08-03
    refresh (units in last 35 days / ~24 selling days, matching the
    Volume/149-selling-days basis of YTD ADU).
  - Seasonal index = (LY Aug-Dec daily rate) / (LY Jan-Jul daily rate),
    from PY Volume and Volume_Prior Year, capped [0.6, 1.8], applied only
    when full-LY volume >= 6 units. Florida's export carries no LY columns,
    so its index is 1.0.

Coverage target (selling days):
  - Primary vendor Oversea:  100 days  (+21 safety days if velocity A)
  - Primary vendor Domestic:  14 days  (+7 safety days if velocity A)

Target inventory = ceil(demand x target days)
Position         = Onhand + OnDock + InTransit + OnOrder
                   (Florida reports one combined pipeline qty -> OnOrder)
Buy qty          = max(0, target - position)

Exclusions: patented items and companywide P-velocity items per the master.
Northeast items join the master on ItemNo; Florida items are partslink-keyed
and join on the master's "Link No_" (rows with a primary vendor preferred).
Rows with no primary vendor default to the Domestic target and are flagged.
"""

import math
import os
import sys
import warnings
from datetime import date

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "output")

CAP_XLSX = os.path.join(HERE, "CAP Raw.xlsx")
FL_XLSX = os.path.join(HERE, "FL Raw.xlsx")
MASTER_CSV = os.path.join(HERE, "KSI_Item_master.csv")
VENDOR_CSV = os.path.join(HERE, "Vendor and Type.csv")

CAP_WAREHOUSES = ["01NJ", "03LF", "05RO", "07BK", "09SJ", "11MH", "13PA", "15MP"]

# Northeast overflow hubs: their combined onhand+intransit can cover buys at
# any other NE DC (netted across DCs - one unit serves one location).
HUBS = ["01NJ", "15MP"]
# 15MP is the top of the supply chain: it feeds the network but is never shown
# as drawing on anyone, so its rows carry no hub availability.
HUB_NO_RECEIVE = ["15MP"]
# Feeder DC -> hub it rolls up into, for the grouped "Warehouse" column.
WH_GROUP = {"07BK": "01NJ", "13PA": "15MP"}
# Feeders carry no inventory of their own: they sell, but ship out of their
# hub. Their demand is folded into the hub's target and they raise no buy
# lines, so a zero inventory position there is expected, not a data fault.
NON_STOCKING = ["07BK", "13PA"]

# Velocity: cumulative share of revenue within each warehouse group.
VEL_A, VEL_B, VEL_C = 0.60, 0.80, 0.95   # A top 60%, B next 20, C next 15, D last 5

SELLING_DAYS_YTD = 149          # Jan 1 - Jul 31 2026 selling days (Volume/ADU in export)
CAL_DAYS_SAME = 212             # calendar days Jan 1 - Jul 31
CAL_DAYS_REMAIN = 153           # calendar days Aug 1 - Dec 31

OVERSEA_DAYS = 100
DOMESTIC_DAYS = 14
SAFETY_A_OVERSEA = 21           # +3 weeks for A items, overseas primary
SAFETY_A_DOMESTIC = 7           # +1 week for A items, domestic primary

SEASONAL_CAP = (0.6, 1.8)
SEASONAL_MIN_LY_UNITS = 6       # need >=6 units full LY to trust a seasonal index

# Recency blend, used per region only when that export's 35-day column is a
# true daily rate (units sold last 35 days / 35). The Northeast "Rolling Avg
# 35 ADU" is avg units per day-with-a-sale (median 1.0) and fails detection.
RECENT_WEIGHT = 0.7             # 70% last-35-day ADU, 30% YTD ADU
RECENT_VALID_MEDIAN = 0.5       # median of positive values must be below this

# Normalized schema every regional loader must produce.
SCHEMA = ["Region", "Warehouse", "ItemNo", "CAP_ItemNum", "Product Desc",
          "Model", "Final Velocity", "Primary Vendor", "Secondary Vendor",
          "Volume", "PY Volume", "Volume_Prior Year", "ADU",
          "Rolling Avg 35 ADU", "Revenue", "Location_Onhand",
          "Location_OnOnDock", "Location_InTransit", "Location_OnOrder"]


MASTER_NOTES = []       # data-quality notes raised while loading the master

# Master fields the model reads. Required ones abort the run if absent;
# optional ones are filled blank and reported, since the master's column set
# has changed between weekly extracts.
MASTER_REQUIRED = ["ItemNo", "Patent", "Companywide veloicty",
                   "Primary Vendor", "Link No_"]
MASTER_OPTIONAL = ["Secondary Vendor"]


def load_master_vendor():
    master = pd.read_csv(MASTER_CSV, encoding="utf-8-sig")
    missing_req = [c for c in MASTER_REQUIRED if c not in master.columns]
    if missing_req:
        raise KeyError(f"KSI_Item_master.csv is missing required column(s) "
                       f"{missing_req}; file has: {list(master.columns)}")
    MASTER_NOTES.clear()
    for c in MASTER_OPTIONAL:
        if c not in master.columns:
            master[c] = ""
            MASTER_NOTES.append(
                f"KSI_Item_master.csv no longer carries a '{c}' column - that "
                f"field is blank in this run's output.")
    master["ItemNo"] = master["ItemNo"].astype(str).str.strip()

    vend = pd.read_csv(VENDOR_CSV, encoding="utf-8-sig")
    vend["V1"] = vend["V1"].astype(str).str.strip()
    vtype = dict(zip(vend["V1"], vend["V1Type"]))
    return master, vtype


def _apply_exclusions(df):
    df = df[df["Patent"].fillna(0) != 1]
    df = df[df["Companywide veloicty"].fillna("") != "P"]
    return df


def _pick(raw, *names):
    """Return the first column present in raw among known aliases —
    the weekly exports have shipped with several naming conventions."""
    for n in names:
        if n in raw.columns:
            return raw[n]
    raise KeyError(f"none of {names} found; file has: {list(raw.columns)}")


def load_northeast(master):
    raw = pd.read_excel(CAP_XLSX)
    df = pd.DataFrame({
        "Warehouse": _pick(raw, "Warehouse", "Location[WarehouseCode]").astype(str).str.strip(),
        "ItemNo": _pick(raw, "ItemNo", "KSI Item[ItemNo]").astype(str).str.strip(),
        "CAP_ItemNum": _pick(raw, "CAP_ItemNum", "KSI Item[CAP_ItemNum]"),
        "Product Desc": _pick(raw, "Product Desc", "[Product_Desc]").fillna(""),
        "Model": _pick(raw, "Model", "[Model]").fillna(""),
        "Final Velocity": _pick(raw, "Final Velocity", "Velocity[Final Velocity]").fillna(""),
        "Volume": _pick(raw, "Volume", "[Volume]"),
        "PY Volume": _pick(raw, "PY Volume", "[PY_Volume]"),
        "Volume_Prior Year": _pick(raw, "Volume_Prior Year", "[Volume_Prior_Year]"),
        "ADU": _pick(raw, "ADU", "[ADU]"),
        "Rolling Avg 35 ADU": _pick(raw, "Rolling Avg 35 ADU", "[Rolling_Avg_35_ADU]"),
        "Revenue": _pick(raw, "Revenue", "[Revenue]"),
        "Location_Onhand": _pick(raw, "Location_Onhand", "[Location_Onhand]"),
        "Location_OnOnDock": _pick(raw, "Location_OnOnDock", "[Location_OnOnDock]"),
        "Location_InTransit": _pick(raw, "Location_InTransit", "[Location_InTransit]"),
        "Location_OnOrder": _pick(raw, "Location_OnOrder", "[Location_OnOrder]"),
    })
    df = df[df["Warehouse"].isin(CAP_WAREHOUSES)].copy()
    for c in ["Volume", "PY Volume", "Volume_Prior Year", "ADU",
              "Rolling Avg 35 ADU", "Revenue", "Location_Onhand",
              "Location_OnOnDock", "Location_InTransit", "Location_OnOrder"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df = df.merge(
        master.drop_duplicates("ItemNo")
        [["ItemNo", "Patent", "Companywide veloicty",
          "Primary Vendor", "Secondary Vendor"]],
        on="ItemNo", how="left",
    )
    df = _apply_exclusions(df)
    df["Region"] = "Northeast"
    return df[SCHEMA]


def master_by_link(master):
    """Master keyed by partslink (Link No_); rows with a vendor win ties."""
    m = master.copy()
    m["_link"] = m["Link No_"].astype(str).str.strip()
    m["_has_v"] = m["Primary Vendor"].notna()
    m = (m.sort_values("_has_v", ascending=False)
         .drop_duplicates("_link"))
    return m.set_index("_link")[["Patent", "Companywide veloicty",
                                 "Primary Vendor", "Secondary Vendor"]]


def load_florida(master):
    raw = pd.read_excel(FL_XLSX, sheet_name="Export")
    df = pd.DataFrame({
        "Warehouse": _pick(raw, "Location[DC]", "DC").astype(str).str.strip(),
        "ItemNo": _pick(raw, "Item[ItemNum]", "ItemNum").astype(str).str.strip(),
        "Product Desc": _pick(raw, "[Part_Desc]", "Part Desc").fillna(""),
        "Final Velocity": _pick(raw, "[Pod_velocity]", "Pod_velocity").fillna(""),
        "Volume": pd.to_numeric(_pick(raw, "[Volume]", "Volume"), errors="coerce").fillna(0),
        "ADU": pd.to_numeric(_pick(raw, "[ADU]", "ADU"), errors="coerce").fillna(0),
        "Rolling Avg 35 ADU": pd.to_numeric(_pick(raw, "[ADU_SO_L35]", "ADU_Last_35"), errors="coerce").fillna(0),
        "Revenue": pd.to_numeric(_pick(raw, "[Revenue]", "Revenue"), errors="coerce").fillna(0),
        "Location_Onhand": pd.to_numeric(_pick(raw, "[Location_Onhand]", "OnHand"), errors="coerce").fillna(0),
        "Location_OnOrder": pd.to_numeric(_pick(raw, "[Qty_InPipeLine]", "InPipeline"), errors="coerce").fillna(0),
        # bool in the old export, 0/1 float in the new one
        "_patented": _pick(raw, "Item[Patented]", "Patented")
                     .map(lambda v: str(v).strip().lower() in ("true", "1", "1.0")),
    })
    df["Region"] = "Florida"
    df["CAP_ItemNum"] = ""
    df["Model"] = ""
    df["PY Volume"] = 0.0
    df["Volume_Prior Year"] = 0.0
    df["Location_OnOnDock"] = 0.0
    df["Location_InTransit"] = 0.0

    mlink = master_by_link(master)
    df = df.join(mlink, on="ItemNo")
    df = df[~df["_patented"]]
    df = _apply_exclusions(df)
    return df[SCHEMA]


def compute(df, vtype):
    """Shared model: vendor typing, per-region demand, targets, buys."""
    df = df.copy()
    df["Primary Vendor"] = df["Primary Vendor"].fillna("").astype(str).str.strip()
    df["Secondary Vendor"] = df["Secondary Vendor"].fillna("").astype(str).str.strip()
    df["Vendor Type"] = df["Primary Vendor"].map(vtype).fillna("")
    df["Vendor Missing"] = df["Primary Vendor"] == ""

    df["ADU"] = df["ADU"].clip(lower=0)

    # --- per-region base demand (recency auto-detection) --------------------
    df["_base"] = df["ADU"]
    modes = {}
    for reg, g in df.groupby("Region"):
        r35 = g["Rolling Avg 35 ADU"].clip(lower=0)
        pos = r35[r35 > 0]
        if len(pos) > 0 and pos.median() < RECENT_VALID_MEDIAN:
            df.loc[g.index, "_base"] = (RECENT_WEIGHT * r35
                                        + (1 - RECENT_WEIGHT) * g["ADU"])
            modes[reg] = (f"blend of {RECENT_WEIGHT:.0%} last-35-day ADU + "
                          f"{1 - RECENT_WEIGHT:.0%} YTD ADU")
        else:
            modes[reg] = ("YTD ADU only (35-day column is not a daily rate "
                          "in this export)")

    ly_same = df["PY Volume"].clip(lower=0)
    ly_full = df["Volume_Prior Year"].clip(lower=0)
    ly_remain = (ly_full - ly_same).clip(lower=0)

    rate_same = ly_same / CAL_DAYS_SAME
    rate_remain = ly_remain / CAL_DAYS_REMAIN
    idx = np.where(
        (ly_full >= SEASONAL_MIN_LY_UNITS) & (rate_same > 0),
        (rate_remain / rate_same.replace(0, np.nan)).fillna(1.0),
        1.0,
    )
    df["Seasonal Index"] = np.clip(idx, *SEASONAL_CAP).round(3)
    df["Own Demand ADU"] = (df["_base"] * df["Seasonal Index"]).round(5)
    df = df.drop(columns="_base")

    # Non-stocking feeders ship out of their hub, so their demand belongs to
    # the hub's target. Fold it in, then zero the feeder's own demand: it
    # keeps its row (for visibility) but raises no buy of its own.
    feeder = df["Warehouse"].isin(NON_STOCKING) & (df["Region"] == "Northeast")
    df["Feeder Demand ADU"] = 0.0
    if feeder.any():
        moved = (df.loc[feeder]
                 .assign(hub=df.loc[feeder, "Warehouse"].map(WH_GROUP))
                 .groupby(["hub", "ItemNo"])["Own Demand ADU"].sum())
        moved = moved[moved > 0]

        # A feeder item the hub does not stock has no row to receive the
        # demand. Create one (zero inventory, zero own demand) so the demand
        # is still covered instead of silently vanishing.
        hub_keys = set(zip(df.loc[df["Warehouse"].isin(HUBS), "Warehouse"],
                           df.loc[df["Warehouse"].isin(HUBS), "ItemNo"]))
        gaps = [k for k in moved.index if k not in hub_keys]
        if gaps:
            src = (df.loc[feeder]
                   .assign(hub=df.loc[feeder, "Warehouse"].map(WH_GROUP))
                   .drop_duplicates(["hub", "ItemNo"])
                   .set_index(["hub", "ItemNo"]))
            new = src.loc[gaps].reset_index()
            new["Warehouse"] = new["hub"]
            new = new.drop(columns="hub")
            for c in ["Volume", "PY Volume", "Volume_Prior Year", "ADU",
                      "Rolling Avg 35 ADU", "Revenue", "Location_Onhand",
                      "Location_OnOnDock", "Location_InTransit",
                      "Location_OnOrder", "Own Demand ADU"]:
                new[c] = 0.0
            new["Seasonal Index"] = 1.0
            df = pd.concat([df, new], ignore_index=True)
            feeder = (df["Warehouse"].isin(NON_STOCKING)
                      & (df["Region"] == "Northeast"))

        key = pd.MultiIndex.from_arrays([df["Warehouse"], df["ItemNo"]])
        is_hub_row = (df["Region"] == "Northeast") & df["Warehouse"].isin(HUBS)
        df.loc[is_hub_row, "Feeder Demand ADU"] = (
            pd.Series(key[is_hub_row].map(moved), index=df.index[is_hub_row])
            .fillna(0.0).values)
    df["Demand ADU"] = (df["Own Demand ADU"]
                        + df["Feeder Demand ADU"]).round(5)
    df.loc[feeder, "Demand ADU"] = 0.0

    # --- warehouse grouping: feeder DCs roll up into their hub --------------
    df["WH Group"] = df["Warehouse"].replace(WH_GROUP)

    # --- velocity: ABC/D by cumulative revenue share within each warehouse
    # group (A = top 60% of revenue, B = next 20%, C = next 15%, D = last 5%).
    # Net returns can make an item's revenue negative; those cannot contribute
    # to a "top X% of revenue" ranking, so rank on revenue floored at zero and
    # classify non-positive items as D.
    rev = (df.groupby(["Region", "WH Group", "ItemNo"], sort=False)["Revenue"]
           .sum().reset_index())
    rev["_r"] = rev["Revenue"].clip(lower=0)
    rev = rev.sort_values(["Region", "WH Group", "_r"],
                          ascending=[True, True, False])
    g = rev.groupby(["Region", "WH Group"], sort=False)["_r"]
    tot = g.transform("sum")
    share = np.where(tot > 0, g.cumsum() / tot.replace(0, np.nan), 1.0)
    rev["Velocity"] = np.select(
        [rev["_r"] <= 0, share <= VEL_A, share <= VEL_B, share <= VEL_C],
        ["D", "A", "B", "C"], default="D")
    df = df.merge(rev[["Region", "WH Group", "ItemNo", "Velocity"]],
                  on=["Region", "WH Group", "ItemNo"], how="left")
    df["Velocity"] = df["Velocity"].fillna("D")

    # --- coverage target ----------------------------------------------------
    # Safety stock follows the COMPUTED per-warehouse velocity (cumulative
    # revenue banding), so the buffer lands only where a SKU actually earns
    # A-class volume at that warehouse.
    oversea = df["Vendor Type"] == "Oversea"
    is_a = df["Velocity"] == "A"

    df["Lead Days"] = np.where(oversea, OVERSEA_DAYS, DOMESTIC_DAYS)
    df["Safety Days"] = np.where(
        is_a, np.where(oversea, SAFETY_A_OVERSEA, SAFETY_A_DOMESTIC), 0
    )
    df["Target Days"] = df["Lead Days"] + df["Safety Days"]

    df["Target Qty"] = np.ceil(df["Demand ADU"] * df["Target Days"]).astype(int)
    df["Position"] = (
        df["Location_Onhand"] + df["Location_OnOnDock"]
        + df["Location_InTransit"] + df["Location_OnOrder"]
    ).astype(int)
    df["Buy Qty"] = (df["Target Qty"] - df["Position"]).clip(lower=0).astype(int)

    # 15MP and 01NJ are the NE overflow hubs. Pool their onhand + intransit +
    # onorder per SKU and expose it on every other NE row, then NET it across
    # those rows: a unit can only cover one warehouse, so allocate the pool
    # highest-demand first.
    is_ne = df["Region"] == "Northeast"
    hub = df[is_ne & (df["Warehouse"].isin(HUBS))]
    # Each hub offers its total inventory: onhand + ondock + intransit +
    # onorder. Locations other than 01NJ and 15MP can draw on both hubs;
    # 01NJ draws on 15MP only (never its own stock, already in its Position);
    # 15MP is the source of supply and draws on nobody.
    stock = (hub["Location_Onhand"] + hub["Location_OnOnDock"]
             + hub["Location_InTransit"] + hub["Location_OnOrder"]).groupby(
                 [hub["ItemNo"], hub["Warehouse"]]).sum()
    per_item = {}
    for (item, wh), qty in stock.items():
        if qty > 0:
            per_item.setdefault(item, {})[wh] = float(qty)

    def sources_for(item, loc):
        """Hub stock this location may draw on: every hub except itself."""
        return {h: q for h, q in per_item.get(item, {}).items() if h != loc}

    # 15MP supplies the network and is never shown drawing on anyone.
    receives = is_ne & (~df["Warehouse"].isin(HUB_NO_RECEIVE))
    df["Hub Avail"] = np.nan
    r_idx = df.index[receives]
    df.loc[r_idx, "Hub Avail"] = [
        sum(sources_for(i, w).values())
        for i, w in zip(df.loc[r_idx, "ItemNo"], df.loc[r_idx, "Warehouse"])]

    # Net the pool: a unit at a hub can only serve one location, so allocate
    # highest demand/day first, drawing from the eligible hubs.
    df["Hub Alloc"] = np.where(receives, 0.0, np.nan)
    cand = df[receives & (df["Buy Qty"] > 0)
              & (df["Hub Avail"] > 0)].sort_values(
        ["ItemNo", "Demand ADU"], ascending=[True, False])
    remaining = {i: dict(h) for i, h in per_item.items()}
    alloc = {}
    for idx, item, loc, need in zip(cand.index, cand["ItemNo"],
                                    cand["Warehouse"], cand["Buy Qty"]):
        pool = remaining.get(item)
        if not pool:
            continue
        took = 0.0
        for h in list(pool):
            if took >= need:
                break
            if h == loc:                 # cannot transfer from itself
                continue
            take = min(pool[h], need - took)
            if take > 0:
                pool[h] -= take
                took += take
                if pool[h] <= 0:
                    del pool[h]
        if took > 0:
            alloc[idx] = took
    if alloc:
        df.loc[list(alloc), "Hub Alloc"] = list(alloc.values())
    df["Net Buy Qty"] = (df["Buy Qty"]
                         - df["Hub Alloc"].fillna(0)).clip(lower=0).astype(int)

    df.attrs["demand_modes"] = modes
    return df


def build_all():
    """Load every available regional export and run the shared model."""
    master, vtype = load_master_vendor()
    frames = [load_northeast(master)]
    if os.path.exists(FL_XLSX):
        frames.append(load_florida(master))
    df = compute(pd.concat(frames, ignore_index=True), vtype)
    return df, master, vtype


def data_quality_warnings(df):
    warns = list(MASTER_NOTES)

    # A Northeast warehouse absent from the export gets no buys at all, which
    # is invisible unless we say so.
    missing = [w for w in CAP_WAREHOUSES
               if w not in set(df.loc[df["Region"] == "Northeast", "Warehouse"])]
    if missing:
        warns.append(
            f"[Northeast] Warehouse(s) {', '.join(missing)} are not present in "
            f"CAP Raw at all - no demand, no inventory, and no buy lines were "
            f"generated for them. Confirm this is intentional, not a dropped "
            f"filter in the export.")

    for (reg, wh), g in df.groupby(["Region", "Warehouse"]):
        if g["Volume"].sum() == 0 and g["Position"].sum() == 0:
            warns.append(
                f"[{reg}] Warehouse {wh}: all volume and inventory are zero in "
                f"this export ({len(g):,} rows) - likely truncated by the source "
                f"system. No buys generated for {wh}; re-export needed."
            )
        elif (g["Volume"].sum() > 0 and g["Position"].sum() == 0
              and wh not in NON_STOCKING):
            # Sales but literally no inventory anywhere: the buy for every row
            # becomes the full target, which overstates the order.
            units = int(g.loc[g["Buy Qty"] > 0, "Buy Qty"].sum())
            warns.append(
                f"[{reg}] Warehouse {wh}: sold {int(g['Volume'].sum()):,} units "
                f"YTD but reports ZERO on-hand, on-dock, in-transit and on-order "
                f"on all {len(g):,} rows. Buy qty there equals the full target "
                f"({units:,} units) and is almost certainly overstated - verify "
                f"the inventory columns for {wh} before ordering."
            )
    for reg, g in df.groupby("Region"):
        n = int((g["Vendor Missing"] & (g["Buy Qty"] > 0)).sum())
        if n:
            warns.append(
                f"[{reg}] {n:,} buy lines have no primary vendor in "
                f"KSI_Item_master - defaulted to the Domestic 14-day target; "
                f"see Exceptions sheet."
            )
    return warns


def display_frame(x):
    """Output naming: Location = physical stocking location, Warehouse = true
    warehouse after rollup, Source Velocity = the export's own velocity."""
    attrs = dict(x.attrs)
    x = x.rename(columns={"Warehouse": "Location",
                          "Final Velocity": "Source Velocity"})
    x["Warehouse"] = x["WH Group"]
    x.attrs = attrs
    return x


def write_excel(df, buys, warns, out_path):
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    df, buys = display_frame(df), display_frame(buys)

    cols = ["Region", "Warehouse", "Location", "ItemNo", "CAP_ItemNum",
            "Product Desc", "Model", "Velocity", "Source Velocity",
            "Primary Vendor", "Vendor Type", "Secondary Vendor", "ADU",
            "Rolling Avg 35 ADU", "Seasonal Index", "Own Demand ADU",
            "Feeder Demand ADU", "Demand ADU", "Lead Days",
            "Safety Days", "Target Days", "Target Qty", "Location_Onhand",
            "Location_OnOnDock", "Location_InTransit", "Location_OnOrder",
            "Position", "Buy Qty", "Hub Avail", "Hub Alloc", "Net Buy Qty",
            "Volume", "PY Volume", "Volume_Prior Year", "Revenue"]

    vend_sum = (buys.groupby(["Region", "Primary Vendor", "Vendor Type"], dropna=False)
                .agg(SKU_Locations=("ItemNo", "size"),
                     Unique_SKUs=("ItemNo", "nunique"),
                     Buy_Units=("Buy Qty", "sum"))
                .reset_index().sort_values("Buy_Units", ascending=False))
    wh_sum = (buys.groupby(["Region", "Warehouse", "Location"])
              .agg(SKU_Locations=("ItemNo", "size"),
                   Buy_Units=("Buy Qty", "sum"),
                   Net_Buy_Units=("Net Buy Qty", "sum"))
              .reset_index().sort_values("Buy_Units", ascending=False))
    vel_sum = (buys.groupby(["Region", "Warehouse", "Velocity"])
               .agg(SKU_Locations=("ItemNo", "size"),
                    Revenue=("Revenue", "sum"),
                    Buy_Units=("Buy Qty", "sum"))
               .reset_index().sort_values(["Region", "Warehouse", "Velocity"]))
    exceptions = df[df["Vendor Missing"] & (df["Buy Qty"] > 0)][cols]

    modes = df.attrs.get("demand_modes", {})
    assumptions = pd.DataFrame({"Assumption": [
        *(f"[{r}] Demand basis: {m}." for r, m in modes.items()),
        "Northeast YTD ADU = Volume / 149 selling days (Jan 1 - Jul 31 2026) as provided in CAP Raw. Florida ADU as provided in FL Raw (source-computed daily rate).",
        "Seasonal index = (LY Aug-Dec daily rate) / (LY Jan-Jul daily rate) from PY Volume and Volume_Prior Year; capped 0.6-1.8; applied only when full-LY volume >= 6 units. FL Raw carries no LY columns, so Florida's index is 1.0.",
        "Non-stocking locations (07BK, 13PA) hold no inventory - they sell but ship out of their hub. Their demand is added to the hub's Demand ADU (see Feeder Demand ADU) and they raise no buy lines of their own; a zero inventory position there is expected.",
        "Location = physical stocking location. Warehouse = true warehouse after rollup: 07BK rolls into 01NJ and 13PA into 15MP; all other locations stand alone. Velocity, the warehouse slicer, and the warehouse chart all use the rolled-up Warehouse.",
        f"Velocity = ABC/D by cumulative share of revenue within each Region + Warehouse group: A = top {VEL_A:.0%} of revenue, B = next {VEL_B-VEL_A:.0%}, C = next {VEL_C-VEL_B:.0%}, D = last {1-VEL_C:.0%} (zero-revenue items are D). This computed Velocity drives the A-item safety stock. The export's companywide letter is kept as Source Velocity for reference only - it is one value per item, identical at every warehouse.",
        "Coverage: Oversea primary = 100 days; Domestic = 14 days. Safety stock (+21 days oversea / +7 domestic) is applied to items whose COMPUTED per-warehouse Velocity is A, so the buffer lands only where the SKU earns A-class revenue at that warehouse. Days = selling days, same basis as ADU.",
        "Hub Avail = hub stock transferable to that row: total inventory (onhand + ondock + intransit + onorder) at the hubs it may draw on. Locations other than 01NJ and 15MP draw on BOTH hubs; 01NJ draws on 15MP only (its own stock is already in its Position); 15MP supplies the network and shows no hub availability. Hub Alloc nets that pool across locations (highest demand/day claims first) so one unit is never counted twice. Net Buy Qty = Buy Qty - Hub Alloc.",
        "Target Qty = ceil(Demand ADU x Target Days). Buy Qty = max(0, Target - (Onhand + OnDock + InTransit + OnOrder)).",
        "Florida reports one combined pipeline quantity (Qty_InPipeLine); it is carried in the Location_OnOrder column, with OnDock/InTransit zero.",
        "Excluded: patented items and companywide P-velocity items per KSI_Item_master (Florida also honors its export's own Patented flag).",
        "Northeast items join the master on ItemNo; Florida items are partslink-keyed and join on the master's Link No_ (rows with a primary vendor preferred).",
        "Rows with no primary vendor use the Domestic 14-day target and appear on the Exceptions sheet - assign vendors to fix.",
        "Negative ADU (net returns) treated as zero demand.",
    ] + ["DATA WARNING: " + w for w in warns]})

    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        buys[cols].to_excel(xw, sheet_name="Buy List", index=False)
        vend_sum.to_excel(xw, sheet_name="Vendor Summary", index=False)
        wh_sum.to_excel(xw, sheet_name="Warehouse Summary", index=False)
        vel_sum.to_excel(xw, sheet_name="Velocity Summary", index=False)
        exceptions.to_excel(xw, sheet_name="Exceptions", index=False)
        assumptions.to_excel(xw, sheet_name="Assumptions", index=False)

        wb = xw.book
        hdr_fill = PatternFill("solid", fgColor="1F4E78")
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.font = Font(name="Arial", bold=True, color="FFFFFF")
                cell.fill = hdr_fill
            ws.freeze_panes = "A2"
            for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                width = min(max(len(str(col[0].value or "")) + 2, 10), 55)
                ws.column_dimensions[get_column_letter(i)].width = width
            ws.auto_filter.ref = ws.dimensions
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Arial")


def write_html(buys, warns, run_date, out_path):
    import json

    b = display_frame(buys)

    def intern(series):
        """Repeated strings -> lookup table + indices (keeps the file small)."""
        s = series.fillna("").astype(str)
        vals = list(pd.unique(s))
        lut = {v: i for i, v in enumerate(vals)}
        return vals, s.map(lut).astype(int)

    descs, i_desc = intern(b["Product Desc"])
    models, i_model = intern(b["Model"])
    vends, i_vend = intern(b["Primary Vendor"])
    types, i_type = intern(b["Vendor Type"])
    vend2, i_vend2 = intern(b["Secondary Vendor"])

    rows = list(zip(
        b["Location"], b["Warehouse"], b["ItemNo"], b["CAP_ItemNum"].fillna(""),
        i_desc, i_model, b["Velocity"], b["Source Velocity"].fillna(""),
        i_vend, i_type, i_vend2,
        b["ADU"].round(4), b["Rolling Avg 35 ADU"].round(4),
        b["Seasonal Index"], b["Demand ADU"].round(4),
        b["Lead Days"].astype(int), b["Safety Days"].astype(int),
        b["Target Days"].astype(int), b["Target Qty"].astype(int),
        b["Location_Onhand"].astype(int), b["Location_OnOnDock"].astype(int),
        b["Location_InTransit"].astype(int), b["Location_OnOrder"].astype(int),
        b["Position"].astype(int), b["Buy Qty"].astype(int),
        b["Hub Avail"].fillna(-1).astype(int),      # -1 = not applicable
        b["Hub Alloc"].fillna(0).astype(int),
        b["Net Buy Qty"].astype(int), b["Region"],
    ))
    payload = json.dumps([list(r) for r in rows], default=str)
    luts = json.dumps({"desc": descs, "model": models, "vend": vends,
                       "type": types, "vend2": vend2})

    warn_html = "".join(f'<div class="warn">&#9888;&#65039; {w}</div>' for w in warns)

    html = HTML_TEMPLATE
    for k, v in {"__DATE__": run_date, "__WARNS__": warn_html,
                 "__LUTS__": luts, "__DATA__": payload}.items():
        html = html.replace(k, v)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>KSI Buy List - __DATE__</title>
<style>
:root{--bg:#12181f;--card:#1a232d;--ink:#e8edf2;--mut:#8fa0b0;--line:#2a3642;--acc:#6aa5d8;--s1:#3987e5;--s2:#d95926;--sx:#898781;color-scheme:dark}
*{box-sizing:border-box}body{margin:0;font:14px/1.5 -apple-system,Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--ink);padding:24px}
h1{font-size:20px;margin:0 0 4px}.sub{color:var(--mut);margin-bottom:14px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:12px;margin-bottom:20px}
.card{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:14px 16px}
.card .v{font-size:22px;font-weight:700}.card .l{color:var(--mut);font-size:12px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:20px}
@media(max-width:800px){.grid2{grid-template-columns:1fr}}
.panel{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:14px 16px;overflow-x:auto}
.panel h2{font-size:14px;margin:0 0 10px}
table{border-collapse:collapse;width:100%;font-size:12px;table-layout:fixed}
th,td{padding:5px 6px;border-bottom:1px solid var(--line);text-align:left;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
th{color:var(--mut);font-size:11px;cursor:pointer;user-select:none}
/* column widths are set inline from the COLS map */
.num{text-align:right;font-variant-numeric:tabular-nums}
.controls{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:10px}
input,select{padding:7px 10px;border:1px solid var(--line);border-radius:8px;background:var(--card);color:var(--ink)}
.pager{display:flex;gap:8px;align-items:center;margin-top:10px;color:var(--mut)}
button{padding:6px 12px;border:1px solid var(--line);border-radius:8px;background:var(--card);color:var(--ink);cursor:pointer}
.badge{display:inline-block;padding:1px 8px;border-radius:10px;font-size:11px;background:var(--acc);color:#0d1319}
.warn{background:#3a3020;border:1px solid #6b5a2a;color:#ffd97a;border-radius:8px;padding:10px 14px;margin-bottom:12px}
.ctrlbar{display:flex;gap:16px;align-items:center;margin:0 0 12px;flex-wrap:wrap}
.lg{display:inline-flex;align-items:center;gap:6px;color:var(--mut);font-size:12px}
.sw{width:10px;height:10px;border-radius:3px;display:inline-block}
.seg{display:inline-flex;border:1px solid var(--line);border-radius:8px;overflow:hidden}
.seg button{border:0;border-radius:0;padding:6px 12px;font-size:12px}
.seg button.on{background:var(--acc);color:#0d1319}
.mlauto{margin-left:auto}
.crow{display:grid;grid-template-columns:118px 1fr 76px;align-items:center;gap:8px;min-height:24px}
.crow .nm{font-size:12px;color:var(--ink);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.crow .val{font-size:12px;color:var(--ink);text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.ghead{font-size:11px;letter-spacing:.06em;text-transform:uppercase;color:var(--mut);margin:12px 0 4px;display:flex;justify-content:space-between}
.ghead:first-child{margin-top:2px}
.track{display:flex;gap:2px;height:16px}
.bseg{height:16px;min-width:1px}
.bseg.end{border-radius:0 4px 4px 0}
#tip{position:fixed;z-index:10;background:var(--card);border:1px solid var(--line);border-radius:8px;padding:8px 10px;font-size:12px;line-height:1.45;pointer-events:none;display:none;box-shadow:0 4px 14px rgba(0,0,0,.18)}
.ok{color:#0ca30c;font-weight:700}
.part{color:#fab219;font-weight:700}
.ms{position:relative;display:inline-block}
.msbtn{display:inline-flex;align-items:center;gap:8px;white-space:nowrap}
.msbtn .cnt{background:var(--acc);color:#0d1319;border-radius:9px;padding:0 6px;font-size:11px;font-weight:700}
.msbtn .car{color:var(--mut);font-size:10px}
.mspanel{display:none;position:absolute;z-index:20;top:calc(100% + 4px);left:0;min-width:190px;max-height:280px;overflow-y:auto;background:var(--card);border:1px solid var(--line);border-radius:8px;padding:6px;box-shadow:0 6px 20px rgba(0,0,0,.35)}
.mspanel.open{display:block}
.msopt{display:flex;align-items:center;gap:8px;padding:5px 8px;border-radius:6px;cursor:pointer;font-size:13px;white-space:nowrap}
.msopt:hover{background:var(--bg)}
.msopt input{margin:0;accent-color:var(--acc)}
.msall{border-bottom:1px solid var(--line);margin-bottom:4px;padding-bottom:7px;color:var(--mut)}
</style></head><body>
<h1>KSI Replenishment Buy List</h1>
<div class="sub">Generated __DATE__ &middot; demand = recency-weighted ADU (where available) &times; seasonal index &middot; oversea 100d / domestic 14d (+A-item safety)</div>
<div class="ctrlbar"><span class="seg" id="rseg"></span></div>
__WARNS__
<div class="cards">
<div class="card"><div class="v" id="cSku"></div><div class="l" id="lblSku">SKU-locations to buy</div></div>
<div class="card"><div class="v" id="cUnits"></div><div class="l" id="lblUnits">Total buy units</div></div>
<div class="card"><div class="v" id="cOs"></div><div class="l">Oversea vendor units</div></div>
<div class="card"><div class="v" id="cDom"></div><div class="l">Domestic vendor units</div></div>
</div>
<div class="ctrlbar">
<span class="lg"><i class="sw" style="background:var(--s1)"></i>Oversea</span>
<span class="lg"><i class="sw" style="background:var(--s2)"></i>Domestic</span>
<span class="lg" id="lgUn" hidden><i class="sw" style="background:var(--sx)"></i>Unassigned vendor</span>
<span class="seg mlauto" id="netseg"><button id="mGross" class="on" title="Full buy quantity required by demand">Gross buy</button><button id="mNet" title="Buy quantity after transferring available 15MP stock (Northeast only)">Net of 15MP transfers</button></span>
<span class="seg"><button id="mUnits" class="on">Units</button><button id="mCtn">Containers (1,300/ctn)</button></span>
</div>
<div class="grid2">
<div class="panel"><h2 id="vh">Top vendors</h2><div id="vchart"></div></div>
<div class="panel"><h2 id="whh">By warehouse</h2><div id="wchart"></div></div>
</div>
<div id="tip"></div>
<div class="panel">
<h2>Buy list detail <span class="badge" id="count"></span></h2>
<div class="controls">
<input id="q" placeholder="Search SKU / CAP # / description / vendor" size="36">
<span class="ms" id="msWh"></span>
<span class="ms" id="msVt"></span>
<span class="ms" id="msVel"></span>
<button id="csv" title="Downloads the rows matching the current filters - always all columns">&#11015; Export CSV (filtered)</button>
<span class="seg mlauto"><button id="cKey" title="Just the decision columns - fits without scrolling">Key columns</button><button id="cAll" class="on" title="Every calculated column (scrolls sideways)">All columns</button></span>
</div>
<table id="tbl"><thead><tr id="thr"></tr></thead><tbody></tbody></table>
<div class="pager"><button id="prev">&laquo; Prev</button><span id="pinfo"></span><button id="next">Next &raquo;</button></div>
</div>
<script>
const DATA=__DATA__,L=__LUTS__;
// column map: i=field index, t=title, n=numeric, f=formatter
const S=v=>v, LU=k=>(v=>L[k][v]||'');
const COLS=[
{h:'Warehouse',i:1,w:76,k:1,t:'True warehouse after rollup: 07BK into 01NJ, 13PA into 15MP'},
{h:'Location',i:0,w:60,k:1,t:'Physical stocking location'},
{h:'SKU',i:2,w:76,k:1},
{h:'CAP Part #',i:3,w:78,k:1},
{h:'Description',i:4,f:LU('desc'),w:200,k:1},
{h:'Model',i:5,f:LU('model'),w:104},
{h:'Vel',i:6,w:36,k:1,t:'ABC/D by cumulative revenue share within the warehouse group (A=top 60%, B=next 20, C=next 15, D=last 5)'},
{h:'Src vel',i:7,w:48,t:"The export's own velocity, kept for reference"},
{h:'Vendor',i:8,f:LU('vend'),w:84,k:1},
{h:'Type',i:9,f:LU('type'),w:60,k:1},
{h:'2nd vendor',i:10,f:LU('vend2'),w:78},
{h:'ADU',i:11,n:1,d:3,w:52,t:'YTD units per selling day'},
{h:'L35 ADU',i:12,n:1,d:3,w:58,t:'Units per selling day over the last 35 days'},
{h:'Seas idx',i:13,n:1,d:3,w:56,t:'Prior-year seasonal multiplier (capped 0.6-1.8)'},
{h:'Dem/day',i:14,n:1,d:3,w:58,k:1,t:'Recency-weighted, seasonally adjusted units per selling day'},
{h:'Lead d',i:15,n:1,w:44,t:'Vendor lead days: oversea 100 / domestic 14'},
{h:'Safety d',i:16,n:1,w:50,t:'A-item safety days: +21 oversea / +7 domestic'},
{h:'Tgt days',i:17,n:1,w:50,t:'Lead + safety days'},
{h:'Target',i:18,n:1,w:52,k:1,t:'Target inventory = demand/day x target days'},
{h:'OnHand',i:19,n:1,w:54},
{h:'OnDock',i:20,n:1,w:54},
{h:'InTrans',i:21,n:1,w:54},
{h:'OnOrder',i:22,n:1,w:54},
{h:"Pos'n",i:23,n:1,w:50,k:1,t:'On hand + on dock + in transit + on order'},
{h:'Buy',i:24,n:1,w:46,k:1,t:'Gross buy = target - position'},
{h:'Hub avail',i:25,n:1,w:74,k:1,t:'Hub inventory (onhand + ondock + intransit + onorder) transferable to this row. Other locations draw on 01NJ + 15MP; 01NJ draws on 15MP only; 15MP shows none. Netted across locations, highest demand claims first. ✓ = fully coverable by transfer; (n) = only n claimable here.'},
{h:'Net buy',i:27,n:1,w:56,k:1,t:'Buy after transferring the claimable hub stock'}];
let allCols=true;   // false = key columns only (fits one screen)
const shown=()=>allCols?COLS:COLS.filter(c=>c.k);
let REGION='',view=[],page=0,PS=100,sortK=24,sortD=-1;
const $=id=>document.getElementById(id);
const REGIONS=[...new Set(DATA.map(r=>r[28]))];
const RDATA=()=>REGION?DATA.filter(r=>r[28]===REGION):DATA;

// region toggle
const rseg=$('rseg');
[['','All regions'],...REGIONS.map(r=>[r,r])].forEach(([val,label])=>{
const b=document.createElement('button');b.textContent=label;b.dataset.r=val;
b.onclick=()=>{REGION=val;[...rseg.children].forEach(x=>x.classList.toggle('on',x.dataset.r===val));refreshFilters();apply();renderCards();renderCharts()};
rseg.appendChild(b)});
rseg.firstChild.classList.add('on');
if(REGIONS.length<2)rseg.parentElement.style.display='none';

// multi-select dropdown: empty selection = all (no filter)
function makeMS(elId,allLabel,noun){
const el=$(elId),sel=new Set();let opts=[];
const btn=document.createElement('button');btn.className='msbtn';
const panel=document.createElement('div');panel.className='mspanel';
el.append(btn,panel);
btn.onclick=e=>{e.stopPropagation();const was=panel.classList.contains('open');
document.querySelectorAll('.mspanel.open').forEach(p=>p.classList.remove('open'));
if(!was)panel.classList.add('open')};
panel.onclick=e=>e.stopPropagation();
function label(){btn.innerHTML=sel.size===0?`${allLabel} <span class="car">&#9662;</span>`
:(sel.size===1?[...sel][0]:`${sel.size} ${noun}`)+` <span class="cnt">${sel.size}</span> <span class="car">&#9662;</span>`}
let allBox=null,boxes=[];
function sync(){if(allBox)allBox.checked=sel.size===0;
boxes.forEach(([o,b])=>{b.checked=sel.has(o)});label()}
function build(){panel.innerHTML='';boxes=[];
const all=document.createElement('label');all.className='msopt msall';
all.innerHTML=`<input type="checkbox"><span>${allLabel}</span>`;
allBox=all.querySelector('input');
allBox.onchange=()=>{sel.clear();sync();apply()};
panel.appendChild(all);
for(const o of opts){const l=document.createElement('label');l.className='msopt';
l.innerHTML=`<input type="checkbox"><span>${o}</span>`;
const b=l.querySelector('input');boxes.push([o,b]);
b.onchange=()=>{b.checked?sel.add(o):sel.delete(o);sync();apply()};
panel.appendChild(l)}
sync()}
return{setOptions(list){opts=list;[...sel].forEach(v=>{if(!opts.includes(v))sel.delete(v)});build()},
match(v){return sel.size===0||sel.has(v)}}}
const MS={wh:makeMS('msWh','All warehouses','warehouses'),
vt:makeMS('msVt','All vendor types','types'),
vel:makeMS('msVel','All velocities','velocities')};
document.addEventListener('click',()=>document.querySelectorAll('.mspanel.open').forEach(p=>p.classList.remove('open')));
document.addEventListener('keydown',e=>{if(e.key==='Escape')document.querySelectorAll('.mspanel.open').forEach(p=>p.classList.remove('open'))});

// headers from the column map (widths inline so table-layout:fixed honors them)
function buildHead(){const cs=shown();
$('thr').innerHTML=cs.map(c=>`<th data-k="${c.i}" class="${c.n?'num':''}" style="width:${c.w}px"${c.t?` title="${c.t}"`:''}>${c.h}</th>`).join('');
$('tbl').style.minWidth=cs.reduce((s,c)=>s+c.w,0)+'px';
document.querySelectorAll('#tbl th[data-k]').forEach(th=>th.onclick=()=>{
const k=+th.dataset.k;sortD=(k===sortK)?-sortD:-1;sortK=k;apply()});
$('cAll').classList.toggle('on',allCols);$('cKey').classList.toggle('on',!allCols)}

function refreshFilters(){const rows=RDATA();
MS.wh.setOptions([...new Set(rows.map(r=>r[1]))].sort());   // true warehouse (post-rollup)
MS.vt.setOptions([...new Set(rows.map(r=>r[9]).map(LU('type')).filter(Boolean))].sort());
MS.vel.setOptions([...new Set(rows.map(r=>r[6]).filter(Boolean))].sort())}
function apply(){const q=$('q').value.toLowerCase(),D=LU('desc'),V=LU('vend'),T=LU('type');
view=RDATA().filter(r=>MS.wh.match(r[1])&&MS.vt.match(T(r[9]))&&MS.vel.match(r[6])
&&(!q||(r[2]+' '+r[3]+' '+D(r[4])+' '+V(r[8])).toLowerCase().includes(q)));
view.sort((a,b)=>{const x=a[sortK],y=b[sortK];return(typeof x==='number'?x-y:String(x).localeCompare(String(y)))*sortD});
page=0;render()}
// hub-availability cell: number + coverage marker + explanation
function hubCell(r){const avail=r[25],alloc=r[26],buy=r[24];
if(avail<0)return['',''];
if(alloc>=buy&&buy>0)return[avail.toLocaleString()+' <span class="ok">&#10003;</span>',
`Transferable hub stock for this SKU: ${avail}; ${alloc} claimable here - covers the full ${buy}-unit buy. Transfer instead of buying.`];
if(alloc>0)return[alloc===avail?`<span class="part">${avail}</span>`:`${avail} <span class="part">(${alloc})</span>`,
alloc===avail?`${avail} transferable - all claimable here, but short of the ${buy}-unit buy. Transfer ${alloc}, still buy ${buy-alloc}.`
:`${avail} transferable, but only ${alloc} is left for this location after higher-demand DCs claim theirs. Transfer ${alloc}, still buy ${buy-alloc}.`];
if(avail>0)return[avail.toLocaleString(),
`${avail} transferable, but it is fully claimed by higher-demand locations for this SKU. No transfer available - buy all ${buy}.`];
return['0',`No hub stock for this SKU - buy all ${buy}.`]}
function render(){const tb=$('tbl').querySelector('tbody');tb.innerHTML='';
view.slice(page*PS,(page+1)*PS).forEach(r=>{const tr=document.createElement('tr');
const [av,avt]=hubCell(r);
tr.innerHTML=shown().map(c=>{
if(c.i===25)return `<td class="num" title="${avt}">${av}</td>`;
let v=c.f?c.f(r[c.i]):r[c.i];
if(c.n)return `<td class="num">${c.d?(+v).toFixed(c.d):(+v).toLocaleString()}</td>`;
const s=v===''||v==null?'':String(v);
return `<td title="${s.replace(/"/g,'&quot;')}">${s}</td>`}).join('');
tb.appendChild(tr)});
$('count').textContent=view.length.toLocaleString()+' rows';
$('pinfo').textContent=`page ${page+1} / ${Math.max(1,Math.ceil(view.length/PS))}`}
$('q').addEventListener('input',apply);
$('prev').onclick=()=>{if(page>0){page--;render()}};
$('next').onclick=()=>{if((page+1)*PS<view.length){page++;render()}};
$('cKey').onclick=()=>{allCols=false;buildHead();render()};
$('cAll').onclick=()=>{allCols=true;buildHead();render()};
buildHead();
$('csv').onclick=()=>{
const hdr=['Region','Warehouse','Location','ItemNo','CAP_ItemNum','Description','Model','Velocity','Source Velocity','Primary Vendor','Vendor Type','Secondary Vendor','ADU','L35 ADU','Seasonal Index','Demand ADU','Lead Days','Safety Days','Target Days','Target Qty','OnHand','OnDock','InTransit','OnOrder','Position','Buy Qty','Hub Avail','Hub Claimable','Net Buy Qty'];
const esc=v=>{v=(v==null?'':String(v));return /[",\n]/.test(v)?'"'+v.replace(/"/g,'""')+'"':v};
const D=LU('desc'),M=LU('model'),V=LU('vend'),T=LU('type'),V2=LU('vend2');
const csv=[hdr.join(',')].concat(view.map(r=>[r[28],r[1],r[0],r[2],r[3],D(r[4]),M(r[5]),r[6],r[7],V(r[8]),T(r[9]),V2(r[10]),
r[11],r[12],r[13],r[14],r[15],r[16],r[17],r[18],r[19],r[20],r[21],r[22],r[23],r[24],
r[25]<0?'':r[25],r[25]<0?'':r[26],r[27]].map(esc).join(','))).join('\r\n');
const a=document.createElement('a');
a.href=URL.createObjectURL(new Blob(['\uFEFF'+csv],{type:'text/csv;charset=utf-8'}));
const reg=(REGION||'all-regions').toLowerCase().replace(/\s+/g,'-');
a.download='buy_list_'+document.title.split(' - ')[1]+'_'+reg+'_'+(view.length===RDATA().length?'all':'filtered')+'.csv';
a.click();URL.revokeObjectURL(a.href)};

// ---- gross vs net-of-15MP-transfers ----
let netMode=false;
// qty(row): gross buy, or buy minus what 15MP can actually cover for that row
const qty=r=>netMode?r[27]:r[24];
const vtype=r=>L.type[r[9]]||'';

// ---- summary cards (region-scoped) ----
function renderCards(){const rows=RDATA();
let u=0,os=0,dom=0;
for(const r of rows){const q=qty(r);u+=q;const t=vtype(r);if(t==='Oversea')os+=q;else if(t==='Domestic')dom+=q}
$('cSku').textContent=(netMode?rows.filter(r=>qty(r)>0).length:rows.length).toLocaleString();
$('cUnits').textContent=u.toLocaleString();
$('cOs').textContent=os.toLocaleString();
$('cDom').textContent=dom.toLocaleString()}

// ---- charts: top vendors & by warehouse, units <-> containers toggle ----
const CTN=1300;let mode='units';
const COLORS={Oversea:'var(--s1)',Domestic:'var(--s2)',Unassigned:'var(--sx)'};
const fmtC=v=>{const c=v/CTN;return c>0&&c<0.05?'<0.1':c.toLocaleString(undefined,{minimumFractionDigits:1,maximumFractionDigits:1})};
const fmt=v=>mode==='units'?v.toLocaleString():fmtC(v);
const both=v=>v.toLocaleString()+' units · '+fmtC(v)+' ctn';
const tip=$('tip');
function tipMove(e){tip.style.left=Math.min(e.clientX+14,innerWidth-tip.offsetWidth-8)+'px';tip.style.top=Math.min(e.clientY+14,innerHeight-tip.offsetHeight-8)+'px'}
function hoverize(el,html){el.addEventListener('mousemove',e=>{tip.innerHTML=html;tip.style.display='block';tipMove(e)});el.addEventListener('mouseleave',()=>tip.style.display='none')}
function vendAgg(){const m=new Map();
for(const r of RDATA()){const q=qty(r);if(!q)continue;const k=(L.vend[r[8]]||'(none)')+'|'+vtype(r);m.set(k,(m.get(k)||0)+q)}
return[...m.entries()].map(([k,v])=>{const[nm,t]=k.split('|');return[nm,t,v]}).sort((a,b)=>b[2]-a[2])}
function whAgg(){const m=new Map();
for(const r of RDATA()){const q=qty(r),t=vtype(r);if(!m.has(r[1]))m.set(r[1],[0,0,0]);const a=m.get(r[1]);
if(t==='Oversea')a[0]+=q;else if(t==='Domestic')a[1]+=q;else a[2]+=q}
return[...m.entries()].map(([w,a])=>[w,...a]).filter(r=>r[1]+r[2]+r[3]>0).sort((a,b)=>(b[1]+b[2]+b[3])-(a[1]+a[2]+a[3]))}
function renderVend(){
const el=$('vchart');el.innerHTML='';
const VDATA=vendAgg();if(!VDATA.length)return;
const max=Math.max(...VDATA.map(r=>r[2]));
$('lgUn').hidden=!VDATA.some(r=>!r[1]);
for(const g of ['Oversea','Domestic','Unassigned']){
const rows=VDATA.filter(r=>(r[1]||'Unassigned')===g);
if(!rows.length)continue;
const tot=rows.reduce((s,r)=>s+r[2],0);
const top=rows.slice(0,8),rest=rows.slice(8),restSum=rest.reduce((s,r)=>s+r[2],0);
const h=document.createElement('div');h.className='ghead';h.innerHTML=`<span>${g}</span><span>${fmt(tot)}</span>`;el.appendChild(h);
const items=top.map(r=>[r[0],r[2]]);if(restSum)items.push([`Other (${rest.length} vendors)`,restSum]);
for(const [nm,v] of items){
const row=document.createElement('div');row.className='crow';
row.innerHTML=`<span class="nm" title="${nm}">${nm}</span><span class="track"><span class="bseg end" style="width:${Math.max(100*v/max,.4).toFixed(2)}%;background:${COLORS[g]}"></span></span><span class="val">${fmt(v)}</span>`;
hoverize(row.querySelector('.bseg'),`<b>${nm}</b><br>${g}<br>${both(v)}`);
el.appendChild(row);}}}
function renderWh(){
const el=$('wchart');el.innerHTML='';
const WDATA=whAgg();if(!WDATA.length)return;
const max=Math.max(...WDATA.map(r=>r[1]+r[2]+r[3]));
for(const [wh,os,dom,un] of WDATA){
const tot=os+dom+un;
const segs=[['Oversea',os],['Domestic',dom],['Unassigned',un]].filter(s=>s[1]>0);
const row=document.createElement('div');row.className='crow';
row.innerHTML=`<span class="nm">${wh}</span><span class="track">${segs.map((s,i)=>`<span class="bseg${i===segs.length-1?' end':''}" style="width:${(100*s[1]/max).toFixed(2)}%;background:${COLORS[s[0]]}"></span>`).join('')}</span><span class="val">${fmt(tot)}</span>`;
row.querySelectorAll('.bseg').forEach((sg,i)=>hoverize(sg,`<b>${wh}</b><br>${segs[i][0]}: ${both(segs[i][1])}<br>Total: ${both(tot)}`));
el.appendChild(row);}}
function renderCharts(){renderVend();renderWh();
const suffix=netMode?' — net of 15MP transfers':'';
$('vh').textContent='Top vendors — '+(mode==='units'?'buy units':'containers (1,300 units each)')+suffix;
$('whh').textContent=(mode==='units'?'Buy units':'Containers')+' by warehouse'+suffix;
$('mUnits').classList.toggle('on',mode==='units');$('mCtn').classList.toggle('on',mode!=='units');
$('mGross').classList.toggle('on',!netMode);$('mNet').classList.toggle('on',netMode);
$('lblUnits').textContent=netMode?'Net buy units (after 15MP transfers)':'Total buy units';
$('lblSku').textContent=netMode?'SKU-locations still to buy':'SKU-locations to buy'}
$('mUnits').onclick=()=>{mode='units';renderCharts()};
$('mCtn').onclick=()=>{mode='ctn';renderCharts()};
$('mGross').onclick=()=>{netMode=false;renderCards();renderCharts()};
$('mNet').onclick=()=>{netMode=true;renderCards();renderCharts()};
refreshFilters();apply();renderCards();renderCharts();
</script></body></html>
"""


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    run_date = date.today().isoformat()

    df, master, vtype = build_all()
    buys = df[df["Buy Qty"] > 0].sort_values("Buy Qty", ascending=False)

    for reg, mode in df.attrs["demand_modes"].items():
        print(f"demand mode [{reg}]: {mode}")
    warns = data_quality_warnings(df)
    for w in warns:
        print(f"WARNING: {w}")

    xlsx_path = os.path.join(OUT_DIR, f"Buy List {run_date}.xlsx")
    write_excel(df, buys, warns, xlsx_path)
    html_path = os.path.join(OUT_DIR, "CAP.html")
    write_html(buys, warns, run_date, html_path)

    # Publishing copies (repo root CAP.html + index.html redirect for Pages)
    # are only written when this folder is the git repo. A shared copy of the
    # folder just gets output/CAP.html to open locally; pass --local to skip
    # the copies even inside the repo.
    publish = os.path.isdir(os.path.join(HERE, ".git")) and "--local" not in sys.argv
    if publish:
        import shutil
        shutil.copyfile(html_path, os.path.join(HERE, "CAP.html"))
        with open(os.path.join(HERE, "index.html"), "w", encoding="utf-8") as f:
            f.write('<!DOCTYPE html><html><head><meta charset="utf-8">'
                    '<meta http-equiv="refresh" content="0; url=CAP.html">'
                    '<title>KSI Buy List</title></head>'
                    '<body><a href="CAP.html">Open the CAP buy list</a></body></html>')

    print(f"rows after filters: {len(df):,}")
    for reg, g in buys.groupby("Region"):
        print(f"[{reg}] buy lines: {len(g):,}  buy units: {int(g['Buy Qty'].sum()):,}"
              f"  (oversea {int(g.loc[g['Vendor Type']=='Oversea','Buy Qty'].sum()):,}"
              f" / domestic {int(g.loc[g['Vendor Type']=='Domestic','Buy Qty'].sum()):,}"
              f" / no-vendor {int(g['Vendor Missing'].sum()):,})")
    print(f"TOTAL buy lines: {len(buys):,}  buy units: {int(buys['Buy Qty'].sum()):,}")
    print(f"wrote: {xlsx_path}")
    print(f"wrote: {html_path}   <- open this in a browser")
    if publish:
        print(f"wrote: {os.path.join(HERE, 'CAP.html')} (+ index.html) "
              f"- commit and push to publish")
    else:
        print("local-only run: no repo files written. Send the report to "
              "whoever owns the publish step.")


if __name__ == "__main__":
    main()
